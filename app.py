from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    jsonify,
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import secrets
from functools import wraps
import pandas as pd
from openpyxl import Workbook

app = Flask(__name__)
app.config["SECRET_KEY"] = secrets.token_hex(16)

# 👉 Use DATABASE_URL from Render if it exists; otherwise, use local SQLite
db_url = os.getenv("DATABASE_URL", "sqlite:///attendance.db")
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://attendance_db_so9v_user:vvqir6Wc3OoNu3uVIGGgW4cKs0TtjPyw@dpg-d3jd5sh5pdvs73ebt0dg-a.oregon-postgres.render.com/attendance_db_so9v")

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False


from models import (
    db,
    User,
    Course,
    Enrollment,
    AttendanceRecord,
    ClassSession,
    Announcement,
    AnnouncementRead,
)
db.init_app(app)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        user = User.query.get(session["user_id"])
        if not user or user.role != "admin":
            flash("Access denied. Admin privileges required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        matric_number = request.form["matric_number"]
        password = request.form["password"]

        user = User.query.filter_by(matric_number=matric_number).first()

        if user and check_password_hash(user.password, password):
            session["user_id"] = user.id
            session["user_role"] = user.role
            session["user_name"] = user.full_name
            flash(f"Welcome, {user.full_name}!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid matric number or password.", "error")

    return render_template("auth/login.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        full_name = request.form["full_name"]
        matric_number = request.form["matric_number"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("auth/signup.html")

        if User.query.filter_by(matric_number=matric_number).first():
            flash("Matric number already exists.", "error")
            return render_template("auth/signup.html")

        user = User(
            full_name=full_name,
            matric_number=matric_number,
            password=generate_password_hash(password),
            role="student",
        )

        db.session.add(user)
        db.session.commit()

        flash("Registration successful! Please login.", "success")
        return redirect(url_for("login"))

    return render_template("auth/signup.html")


@app.route("/dashboard")
@login_required
def dashboard():
    user = User.query.get(session["user_id"])

    if user.role == "admin":
        return redirect(url_for("admin_dashboard"))

    # Student dashboard data
    enrollments = Enrollment.query.filter_by(user_id=user.id).all()
    courses_count = len(enrollments)

    # Calculate average attendance
    total_classes = 0
    attended_classes = 0

    for enrollment in enrollments:
        course_records = AttendanceRecord.query.filter_by(
            user_id=user.id, course_id=enrollment.course_id
        ).all()

        total_classes += len(course_records)
        attended_classes += len([r for r in course_records if r.status == "present"])

    avg_attendance = round(
        (attended_classes / total_classes * 100) if total_classes > 0 else 0, 1
    )

    # Get upcoming classes
    today = datetime.now().date()
    upcoming_sessions = (
        ClassSession.query.filter(
            ClassSession.date >= today,
            ClassSession.course_id.in_([e.course_id for e in enrollments]),
        )
        .order_by(ClassSession.date, ClassSession.start_time)
        .limit(5)
        .all()
    )

    return render_template(
        "student/dashboard.html",
        user=user,
        courses_count=courses_count,
        avg_attendance=avg_attendance,
        upcoming_sessions=upcoming_sessions,
    )


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    user = User.query.get(session["user_id"])

    # Admin dashboard stats
    total_students = User.query.filter_by(role="student").count()
    total_courses = Course.query.count()

    # Today's classes
    today = datetime.now().date()
    today_sessions = ClassSession.query.filter_by(date=today).all()

    # Recent attendance
    recent_records = (
        AttendanceRecord.query.order_by(AttendanceRecord.timestamp.desc())
        .limit(10)
        .all()
    )

    return render_template(
        "admin/dashboard.html",
        user=user,
        total_students=total_students,
        total_courses=total_courses,
        today_sessions=today_sessions,
        recent_records=recent_records,
    )


@app.route("/courses")
@login_required
def courses():
    user = User.query.get(session["user_id"])

    if user.role == "admin":
        courses = Course.query.all()
        return render_template("admin/courses.html", courses=courses, user=user)

    # Student courses
    enrollments = Enrollment.query.filter_by(user_id=user.id).all()
    enrolled_courses = [Course.query.get(e.course_id) for e in enrollments]

    # Today's courses
    today = datetime.now().date()
    today_sessions = ClassSession.query.filter(
        ClassSession.date == today,
        ClassSession.course_id.in_([c.id for c in enrolled_courses]),
    ).all()

    return render_template(
        "student/courses.html",
        user=user,
        enrolled_courses=enrolled_courses,
        today_sessions=today_sessions,
    )


@app.route("/announcements")
@login_required
def announcements():
    user = User.query.get(session["user_id"])
    return render_template("student/announcements.html", user=user)


@app.route("/settings")
@login_required
def settings():
    user = User.query.get(session["user_id"])
    return render_template("student/settings.html", user=user)


@app.route("/course/<int:course_id>")
@login_required
def course_detail(course_id):
    user = User.query.get(session["user_id"])

    # Check if user is enrolled in this course
    enrollment = Enrollment.query.filter_by(
        user_id=user.id, course_id=course_id
    ).first_or_404()

    course = Course.query.get_or_404(course_id)

    return render_template(
        "student/course_detail.html", user=user, course=course, enrollment=enrollment
    )


@app.route("/api/join-course", methods=["POST"])
@login_required
def api_join_course():
    try:
        data = request.get_json()
        course_code = data.get("course_code", "").strip().upper()

        if not course_code:
            return jsonify({"success": False, "message": "Course code is required"})

        course = Course.query.filter_by(join_code=course_code).first()
        if not course:
            return jsonify({"success": False, "message": "Invalid course code"})

        # Check if already enrolled
        existing = Enrollment.query.filter_by(
            user_id=session["user_id"], course_id=course.id
        ).first()
        if existing:
            return jsonify(
                {"success": False, "message": "Already enrolled in this course"}
            )

        # Create enrollment
        enrollment = Enrollment(user_id=session["user_id"], course_id=course.id)
        db.session.add(enrollment)
        db.session.commit()

        return jsonify({"success": True, "message": "Successfully joined course"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/mark-attendance", methods=["POST"])
@login_required
def api_mark_attendance():
    try:
        data = request.get_json()
        session_id = data.get("session_id")

        if not session_id:
            return jsonify({"success": False, "message": "Session ID is required"})

        class_session = ClassSession.query.get(session_id)
        if not class_session:
            return jsonify({"success": False, "message": "Class session not found"})

        # Check if session is active
        if not class_session.is_active():
            return jsonify(
                {"success": False, "message": "Class session is not currently active"}
            )

        # Check if student is enrolled in the course
        enrollment = Enrollment.query.filter_by(
            user_id=session["user_id"], course_id=class_session.course_id
        ).first()

        if not enrollment:
            return jsonify({"success": False, "message": "Not enrolled in this course"})

        existing_record = AttendanceRecord.query.filter_by(
            user_id=session["user_id"], class_session_id=session_id
        ).first()

        if existing_record:
            return jsonify(
                {
                    "success": False,
                    "message": "Attendance already marked for this session",
                }
            )

        attendance = AttendanceRecord(
            user_id=session["user_id"],
            course_id=class_session.course_id,
            class_session_id=session_id,
            status="present",
            marked_by="student",
        )

        db.session.add(attendance)
        db.session.commit()

        return jsonify({"success": True, "message": "Attendance marked successfully"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


# Admin routes
@app.route("/admin/courses")
@admin_required
def admin_courses():
    courses = Course.query.all()
    return render_template("admin/courses.html", courses=courses)


@app.route("/admin/students")
@admin_required
def admin_students():
    students = User.query.filter_by(role="student").all()
    return render_template("admin/students.html", students=students)


@app.route("/admin/schedule")
@admin_required
def admin_schedule():
    sessions = ClassSession.query.order_by(
        ClassSession.date, ClassSession.start_time
    ).all()
    return render_template("admin/schedule.html", sessions=sessions)


@app.route("/admin/attendance")
@admin_required
def admin_attendance():
    records = (
        AttendanceRecord.query.order_by(AttendanceRecord.timestamp.desc())
        .limit(100)
        .all()
    )
    return render_template("admin/attendance.html", records=records)


@app.route("/admin/exports")
@admin_required
def admin_exports():
    return render_template("admin/exports.html")


# API routes
@app.route("/api/courses")
@login_required
def api_courses():
    courses = Course.query.all()
    return jsonify(
        {
            "success": True,
            "courses": [
                {
                    "id": c.id,
                    "course_code": c.course_code,
                    "course_title": c.course_title,
                    "lecturer_name": c.lecturer_name,
                    "join_code": c.join_code,
                }
                for c in courses
            ],
        }
    )


@app.route("/api/admin/create-course", methods=["POST"])
@admin_required
def api_admin_create_course():
    try:
        data = request.get_json()

        # Check if course code already exists
        existing_course = Course.query.filter_by(
            course_code=data["course_code"]
        ).first()
        if existing_course:
            return jsonify({"success": False, "message": "Course code already exists"})

        course = Course(
            course_code=data["course_code"].upper(),
            course_title=data["course_title"],
            lecturer_name=data["lecturer_name"],
            description=data.get("description", ""),
        )

        db.session.add(course)
        db.session.commit()

        return jsonify(
            {
                "success": True,
                "message": "Course created successfully",
                "join_code": course.join_code,
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/admin/schedule-class", methods=["POST"])
@admin_required
def api_admin_schedule_class():
    try:
        data = request.get_json()

        session = ClassSession(
            course_id=data["course_id"],
            date=datetime.strptime(data["date"], "%Y-%m-%d").date(),
            start_time=datetime.strptime(data["start_time"], "%H:%M").time(),
            end_time=datetime.strptime(data["end_time"], "%H:%M").time(),
            location=data.get("location", ""),
            status="scheduled",
        )

        db.session.add(session)
        db.session.commit()

        return jsonify({"success": True, "message": "Class scheduled successfully"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/admin/export-attendance")
@admin_required
def api_admin_export_attendance():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Get filter parameters
        course_id = request.args.get("course_id")
        student_id = request.args.get("student_id")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        status_filter = request.args.get("status")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Headers
        headers = [
            "Date",
            "Course Code",
            "Course Title",
            "Student Name",
            "Matric Number",
            "Status",
            "Time Marked",
        ]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Build query with filters
        query = (
            db.session.query(AttendanceRecord)
            .join(User)
            .join(Course)
            .join(ClassSession)
        )

        # Apply filters
        if course_id:
            query = query.filter(AttendanceRecord.course_id == course_id)
        if student_id:
            query = query.filter(AttendanceRecord.user_id == student_id)
        if start_date:
            query = query.filter(
                ClassSession.date >= datetime.strptime(start_date, "%Y-%m-%d").date()
            )
        if end_date:
            query = query.filter(
                ClassSession.date <= datetime.strptime(end_date, "%Y-%m-%d").date()
            )
        if status_filter:
            query = query.filter(AttendanceRecord.status == status_filter)

        records = query.order_by(ClassSession.date.desc()).all()

        # Add data rows
        for record in records:
            ws.append(
                [
                    record.class_session.date.strftime("%Y-%m-%d"),
                    record.course.course_code,
                    record.course.course_title,
                    record.user.full_name,
                    record.user.matric_number,
                    record.status.upper(),
                    record.timestamp.strftime("%H:%M:%S"),
                ]
            )

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import Response

        filename = "attendance_export"
        if course_id:
            course = Course.query.get(course_id)
            filename += f"_{course.course_code}" if course else ""
        if start_date or end_date:
            filename += f"_{start_date or 'start'}_to_{end_date or 'end'}"
        filename += ".xlsx"

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/class-sessions", methods=["GET", "POST"])
@admin_required
def api_manage_class_sessions():
    if request.method == "GET":
        try:
            sessions = ClassSession.query.order_by(
                ClassSession.date.desc(), ClassSession.start_time
            ).all()

            sessions_data = []
            for session in sessions:
                sessions_data.append(
                    {
                        "id": session.id,
                        "course_id": session.course_id,
                        "course_code": session.course.course_code,
                        "course_title": session.course.course_title,
                        "date": session.date.isoformat(),
                        "start_time": session.start_time.strftime("%H:%M"),
                        "end_time": session.end_time.strftime("%H:%M"),
                        "location": session.location,
                        "status": session.status,
                    }
                )

            return jsonify({"success": True, "sessions": sessions_data})
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "POST":
        try:
            data = request.get_json()

            session = ClassSession(
                course_id=data["course_id"],
                date=datetime.strptime(data["date"], "%Y-%m-%d").date(),
                start_time=datetime.strptime(data["start_time"], "%H:%M").time(),
                end_time=datetime.strptime(data["end_time"], "%H:%M").time(),
                location=data.get("location", ""),
                status="scheduled",
            )

            db.session.add(session)
            db.session.commit()

            return jsonify(
                {"success": True, "message": "Class session created successfully"}
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})


@app.route("/api/class-sessions/<int:session_id>", methods=["GET", "PUT", "DELETE"])
@admin_required
def api_manage_class_session(session_id):
    session_obj = ClassSession.query.get_or_404(session_id)

    if request.method == "GET":
        try:
            return jsonify(
                {
                    "success": True,
                    "session": {
                        "id": session_obj.id,
                        "course_id": session_obj.course_id,
                        "course_code": session_obj.course.course_code,
                        "course_title": session_obj.course.course_title,
                        "date": session_obj.date.isoformat(),
                        "start_time": session_obj.start_time.strftime("%H:%M"),
                        "end_time": session_obj.end_time.strftime("%H:%M"),
                        "location": session_obj.location,
                        "status": session_obj.status,
                    },
                }
            )
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "PUT":
        try:
            data = request.get_json()

            session_obj.course_id = data.get("course_id", session_obj.course_id)
            session_obj.date = (
                datetime.strptime(data["date"], "%Y-%m-%d").date()
                if "date" in data
                else session_obj.date
            )
            session_obj.start_time = (
                datetime.strptime(data["start_time"], "%H:%M").time()
                if "start_time" in data
                else session_obj.start_time
            )
            session_obj.end_time = (
                datetime.strptime(data["end_time"], "%H:%M").time()
                if "end_time" in data
                else session_obj.end_time
            )
            session_obj.location = data.get("location", session_obj.location)
            session_obj.status = data.get("status", session_obj.status)

            db.session.commit()
            return jsonify(
                {"success": True, "message": "Class session updated successfully"}
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "DELETE":
        try:
            db.session.delete(session_obj)
            db.session.commit()
            return jsonify(
                {"success": True, "message": "Class session deleted successfully"}
            )
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})


@app.route("/api/admin/session/<int:session_id>/live-attendance")
@admin_required
def api_admin_live_attendance(session_id):
    try:
        session = ClassSession.query.get_or_404(session_id)

        # Get all enrolled students
        enrollments = Enrollment.query.filter_by(course_id=session.course_id).all()
        enrolled_students = [e.user for e in enrollments]

        # Get attendance records for this session
        attendance_records = AttendanceRecord.query.filter_by(
            class_session_id=session_id
        ).all()
        present_students = {
            r.user_id: r for r in attendance_records if r.status == "present"
        }

        students_data = []
        for student in enrolled_students:
            status = "present" if student.id in present_students else "absent"
            students_data.append(
                {
                    "id": student.id,
                    "name": student.full_name,
                    "matric_number": student.matric_number,
                    "status": status,
                }
            )

        return jsonify(
            {
                "success": True,
                "course": {
                    "course_code": session.course.course_code,
                    "course_title": session.course.course_title,
                },
                "total_enrolled": len(enrolled_students),
                "present_count": len(present_students),
                "students": students_data,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/course/<int:course_id>/details")
@login_required
def api_course_details(course_id):
    try:
        course = Course.query.get_or_404(course_id)
        user_id = session["user_id"]

        # Get attendance records for this user and course
        records = (
            AttendanceRecord.query.filter_by(user_id=user_id, course_id=course_id)
            .join(ClassSession)
            .order_by(ClassSession.date.desc())
            .all()
        )

        attendance_data = []
        for record in records:
            attendance_data.append(
                {
                    "date": record.class_session.date.strftime("%Y-%m-%d"),
                    "status": record.status,
                    "timestamp": record.timestamp.strftime("%H:%M"),
                }
            )

        total_classes = len(records)
        classes_attended = len([r for r in records if r.status == "present"])
        attendance_percentage = round(
            (classes_attended / total_classes * 100) if total_classes > 0 else 0, 1
        )

        return jsonify(
            {
                "success": True,
                "course": {
                    "course_code": course.course_code,
                    "course_title": course.course_title,
                    "lecturer_name": course.lecturer_name,
                },
                "total_classes": total_classes,
                "classes_attended": classes_attended,
                "attendance_percentage": attendance_percentage,
                "attendance": attendance_data,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/course/<int:course_id>/weekly-attendance")
@login_required
def api_course_weekly_attendance(course_id):
    try:
        course = Course.query.get_or_404(course_id)
        user_id = session["user_id"]

        # Check if user is enrolled
        enrollment = Enrollment.query.filter_by(
            user_id=user_id, course_id=course_id
        ).first()
        if not enrollment:
            return jsonify({"success": False, "message": "Not enrolled in this course"})

        weekly_data = course.get_weekly_attendance(user_id)

        return jsonify(
            {
                "success": True,
                "course": {
                    "id": course.id,
                    "course_code": course.course_code,
                    "course_title": course.course_title,
                    "lecturer_name": course.lecturer_name,
                },
                "weekly_attendance": weekly_data,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/courses/<int:course_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def api_manage_course(course_id):
    course = Course.query.get_or_404(course_id)

    if request.method == "GET":
        try:
            return jsonify(
                {
                    "success": True,
                    "course": {
                        "id": course.id,
                        "course_code": course.course_code,
                        "course_title": course.course_title,
                        "lecturer_name": course.lecturer_name,
                        "description": course.description,
                        "join_code": course.join_code,
                        "created_date": course.created_at.isoformat(),
                        "enrollment_count": course.get_enrollment_count(),
                        "class_count": len(course.class_sessions),
                    },
                }
            )
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "PUT":
        if session.get("user_role") != "admin":
            return jsonify({"success": False, "message": "Admin access required"})

        try:
            data = request.get_json()
            course.course_code = data.get("course_code", course.course_code)
            course.course_title = data.get("course_title", course.course_title)
            course.lecturer_name = data.get("lecturer_name", course.lecturer_name)
            course.description = data.get("description", course.description)

            db.session.commit()
            return jsonify({"success": True, "message": "Course updated successfully"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "DELETE":
        if session.get("user_role") != "admin":
            return jsonify({"success": False, "message": "Admin access required"})

        try:
            db.session.delete(course)
            db.session.commit()
            return jsonify({"success": True, "message": "Course deleted successfully"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})


@app.route("/api/students", methods=["GET", "POST"])
@admin_required
def api_manage_students():
    if request.method == "GET":
        try:
            course_id = request.args.get("course")
            if course_id:
                # Filter students by course enrollment
                enrollments = Enrollment.query.filter_by(course_id=course_id).all()
                students = [e.user for e in enrollments]
            else:
                students = User.query.filter_by(role="student").all()

            students_data = []
            for student in students:
                students_data.append(
                    {
                        "id": student.id,
                        "full_name": student.full_name,
                        "matric_number": student.matric_number,
                        "created_at": student.created_at.isoformat(),
                        "enrollment_count": len(student.enrollments),
                    }
                )

            return jsonify({"success": True, "students": students_data})
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "POST":
        try:
            data = request.get_json()

            # Check if matric number already exists
            if User.query.filter_by(matric_number=data["matric_number"]).first():
                return jsonify(
                    {"success": False, "message": "Matric number already exists"}
                )

            student = User(
                full_name=data["full_name"],
                matric_number=data["matric_number"],
                password=generate_password_hash(data["password"]),
                role="student",
            )

            db.session.add(student)
            db.session.flush()  # Get the student ID

            # Enroll in courses if specified
            if "courses" in data and data["courses"]:
                for course_id in data["courses"]:
                    enrollment = Enrollment(
                        user_id=student.id, course_id=int(course_id)
                    )
                    db.session.add(enrollment)

            db.session.commit()
            return jsonify({"success": True, "message": "Student added successfully"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})


@app.route("/api/students/<int:student_id>", methods=["GET", "PUT", "DELETE"])
@admin_required
def api_manage_student(student_id):
    student = User.query.filter_by(id=student_id, role="student").first_or_404()

    if request.method == "GET":
        try:
            enrollments = [
                {
                    "course_id": e.course.id,
                    "course_code": e.course.course_code,
                    "course_title": e.course.course_title,
                    "attendance_rate": e.course.get_attendance_percentage(student.id),
                }
                for e in student.enrollments
            ]

            return jsonify(
                {
                    "success": True,
                    "student": {
                        "id": student.id,
                        "full_name": student.full_name,
                        "matric_number": student.matric_number,
                        "created_at": student.created_at.isoformat(),
                        "enrollment_count": len(student.enrollments),
                        "enrollments": enrollments,
                    },
                }
            )
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "PUT":
        try:
            data = request.get_json()
            student.full_name = data.get("full_name", student.full_name)
            student.matric_number = data.get("matric_number", student.matric_number)

            if "password" in data and data["password"]:
                student.password = generate_password_hash(data["password"])

            db.session.commit()
            return jsonify({"success": True, "message": "Student updated successfully"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})

    elif request.method == "DELETE":
        try:
            db.session.delete(student)
            db.session.commit()
            return jsonify({"success": True, "message": "Student deleted successfully"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)})


@app.route("/api/students/<int:student_id>/enrollments")
@admin_required
def api_student_enrollments(student_id):
    try:
        student = User.query.filter_by(id=student_id, role="student").first_or_404()

        # Get current enrollments
        enrollments = [
            {
                "course_id": e.course.id,
                "course_code": e.course.course_code,
                "course_title": e.course.course_title,
            }
            for e in student.enrollments
        ]

        # Get available courses (not enrolled)
        enrolled_course_ids = [e.course_id for e in student.enrollments]
        available_courses = Course.query.filter(
            ~Course.id.in_(enrolled_course_ids)
        ).all()

        available_data = [
            {"id": c.id, "course_code": c.course_code, "course_title": c.course_title}
            for c in available_courses
        ]

        return jsonify(
            {
                "success": True,
                "student": {
                    "id": student.id,
                    "full_name": student.full_name,
                    "enrollments": enrollments,
                },
                "available_courses": available_data,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/students/<int:student_id>/enroll", methods=["POST"])
@admin_required
def api_enroll_student(student_id):
    try:
        data = request.get_json()
        course_id = data.get("course_id")

        # Check if already enrolled
        existing = Enrollment.query.filter_by(
            user_id=student_id, course_id=course_id
        ).first()
        if existing:
            return jsonify({"success": False, "message": "Already enrolled"})

        enrollment = Enrollment(user_id=student_id, course_id=course_id)
        db.session.add(enrollment)
        db.session.commit()

        return jsonify({"success": True, "message": "Student enrolled successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/students/<int:student_id>/unenroll", methods=["POST"])
@admin_required
def api_unenroll_student(student_id):
    try:
        data = request.get_json()
        course_id = data.get("course_id")

        enrollment = Enrollment.query.filter_by(
            user_id=student_id, course_id=course_id
        ).first_or_404()

        db.session.delete(enrollment)
        db.session.commit()

        return jsonify({"success": True, "message": "Student unenrolled successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/students/<int:student_id>/password")
@admin_required
def api_student_password(student_id):
    """Get student password for admin viewing (placeholder for security)"""
    try:
        student = User.query.filter_by(id=student_id, role="student").first_or_404()
        # For security reasons, we return a placeholder that suggests the original password
        # In a real-world scenario, you'd want to store original passwords in a secure vault
        # or generate temporary passwords. For this demo, we'll return a standardized password.
        return jsonify(
            {
                "success": True,
                "password": f"student{student.matric_number}",  # Standardized format
            }
        )
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/enrollments", methods=["POST"])
@admin_required
def api_create_enrollment():
    """Create a new enrollment"""
    try:
        data = request.get_json()
        student_id = data.get("student_id")
        course_id = data.get("course_id")

        # Check if already enrolled
        existing = Enrollment.query.filter_by(
            user_id=student_id, course_id=course_id
        ).first()
        if existing:
            return jsonify({"success": False, "message": "Already enrolled"})

        enrollment = Enrollment(user_id=student_id, course_id=course_id)
        db.session.add(enrollment)
        db.session.commit()

        return jsonify({"success": True, "message": "Student enrolled successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/enrollments/<int:student_id>/<int:course_id>", methods=["DELETE"])
@admin_required
def api_delete_enrollment(student_id, course_id):
    """Delete an enrollment"""
    try:
        enrollment = Enrollment.query.filter_by(
            user_id=student_id, course_id=course_id
        ).first_or_404()

        db.session.delete(enrollment)
        db.session.commit()

        return jsonify({"success": True, "message": "Student unenrolled successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    try:
        data = request.get_json()
        current_password = data.get("current_password")
        new_password = data.get("new_password")

        if not current_password or not new_password:
            return jsonify(
                {"success": False, "message": "Current and new passwords are required"}
            )

        user = User.query.get(session["user_id"])

        if not check_password_hash(user.password, current_password):
            return jsonify(
                {"success": False, "message": "Current password is incorrect"}
            )

        user.password = generate_password_hash(new_password)
        db.session.commit()

        return jsonify({"success": True, "message": "Password changed successfully"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/announcements")
@login_required
def api_announcements():
    try:
        user_id = session["user_id"]
        user = User.query.get(user_id)

        # Get announcements visible to this user
        if user.role == "admin":
            announcements = Announcement.query.order_by(
                Announcement.created_at.desc()
            ).all()
        else:
            # Students see general announcements and course-specific ones
            enrolled_course_ids = [e.course_id for e in user.enrollments]
            announcements = (
                Announcement.query.filter(
                    db.or_(
                        Announcement.course_id.is_(None),  # General announcements
                        Announcement.course_id.in_(
                            enrolled_course_ids
                        ),  # Course-specific
                    )
                )
                .order_by(Announcement.created_at.desc())
                .all()
            )

        announcements_data = []
        for announcement in announcements:
            announcements_data.append(
                {
                    "id": announcement.id,
                    "title": announcement.title,
                    "content": announcement.content,
                    "priority": announcement.priority,
                    "created_at": announcement.created_at.isoformat(),
                    "author": announcement.author.full_name,
                    "course_code": (
                        announcement.course.course_code if announcement.course else None
                    ),
                    "is_read": announcement.is_read_by(user_id),
                }
            )

        unread_count = len([a for a in announcements_data if not a["is_read"]])

        return jsonify(
            {
                "success": True,
                "announcements": announcements_data,
                "unread_count": unread_count,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/announcements", methods=["POST"])
@admin_required
def api_create_announcement():
    try:
        data = request.get_json()

        announcement = Announcement(
            title=data["title"],
            content=data["content"],
            author_id=session["user_id"],
            course_id=data.get("course_id") if data.get("course_id") else None,
            priority=data.get("priority", "normal"),
        )

        db.session.add(announcement)
        db.session.commit()

        return jsonify(
            {"success": True, "message": "Announcement created successfully"}
        )
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/announcements/<int:announcement_id>/read", methods=["POST"])
@login_required
def api_mark_announcement_read(announcement_id):
    try:
        user_id = session["user_id"]

        # Check if already marked as read
        existing = AnnouncementRead.query.filter_by(
            announcement_id=announcement_id, user_id=user_id
        ).first()

        if not existing:
            read_record = AnnouncementRead(
                announcement_id=announcement_id, user_id=user_id
            )
            db.session.add(read_record)
            db.session.commit()

        return jsonify({"success": True, "message": "Announcement marked as read"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/api/announcements/mark-all-read", methods=["POST"])
@login_required
def api_mark_all_announcements_read():
    try:
        user_id = session["user_id"]
        user = User.query.get(user_id)

        # Get all unread announcements for this user
        if user.role == "admin":
            announcements = Announcement.query.all()
        else:
            enrolled_course_ids = [e.course_id for e in user.enrollments]
            announcements = Announcement.query.filter(
                db.or_(
                    Announcement.course_id.is_(None),
                    Announcement.course_id.in_(enrolled_course_ids),
                )
            ).all()

        # Mark all as read
        for announcement in announcements:
            if not announcement.is_read_by(user_id):
                read_record = AnnouncementRead(
                    announcement_id=announcement.id, user_id=user_id
                )
                db.session.add(read_record)

        db.session.commit()
        return jsonify({"success": True, "message": "All announcements marked as read"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("index"))


def migrate_database_if_needed():
    """Run database migrations if needed"""
    import sqlite3
    import os

    db_path = "attendance.db"

    if not os.path.exists(db_path):
        print("No existing database found, will create new one")
        return

    print("Checking for database updates...")


if __name__ == "__main__":
    with app.app_context():
        try:
            print("Initializing database...")

            # Check if we need migrations first
            migrate_database_if_needed()

            # Create all tables
            db.create_all()
            print("[SUCCESS] Database tables created successfully")

            # Create admin user if it doesn't exist
            admin = User.query.filter_by(matric_number="admin").first()
            if not admin:
                admin = User(
                    full_name="System Administrator",
                    matric_number="admin",
                    password=generate_password_hash("admin123"),
                    role="admin",
                )
                db.session.add(admin)
                db.session.commit()
                print(
                    "[SUCCESS] Admin user created: matric_number='admin', password='admin123'"
                )
            else:
                print("[INFO] Admin user already exists")

        except Exception as e:
            print(f"[ERROR] Database initialization error: {e}")
            print("If this persists, try deleting 'attendance.db' and restarting")
            raise

    print("\n=== CS Present Attendance System ===")
    print("Admin Panel: http://localhost:5000 (login: admin/admin123)")
    print("Student Portal: http://localhost:5000 (students create accounts)")
    print("\nFeatures included:")
    print("  * Course Management with CRUD operations")
    print("  * Student password management")
    print("  * Real-time announcements system")
    print("  * Weekly attendance tracking")
    print("  * Filtered data exports")
    print("  * Mobile-responsive design")
    print("\nPress Ctrl+C to stop the server")

    app.run(host="0.0.0.0", port=5000)
